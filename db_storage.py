"""
Database storage for Railway deployment
Uses PostgreSQL when DATABASE_URL is available, falls back to Excel locally
"""

import os
import logging
from datetime import datetime
from typing import Optional

logger = logging.getLogger(__name__)

DATABASE_URL = os.environ.get("DATABASE_URL")

if DATABASE_URL:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    
    def get_connection():
        return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    
    def init_db():
        """Create tables if they don't exist"""
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS employees (
                id SERIAL PRIMARY KEY,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                action VARCHAR(50),
                name VARCHAR(255),
                wallet VARCHAR(255),
                telegram VARCHAR(255),
                language VARCHAR(10)
            )
        """)
        conn.commit()
        cur.close()
        conn.close()
        logger.info("Database initialized")
    
    def save_employee(action: str, name: str, wallet: str, telegram: str, language: str):
        """Save employee data to PostgreSQL"""
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            """INSERT INTO employees (action, name, wallet, telegram, language) 
               VALUES (%s, %s, %s, %s, %s)""",
            (action, name, wallet, telegram, language)
        )
        conn.commit()
        cur.close()
        conn.close()
        logger.info(f"Saved to DB: {name}")
        return True
    
    def get_all_employees():
        """Get all employees from PostgreSQL"""
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM employees ORDER BY timestamp DESC")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return rows

else:
    # Fallback to Excel for local development
    from openpyxl import Workbook, load_workbook
    
    EXCEL_FILE = os.path.join(os.path.dirname(__file__), "data.xlsx")
    
    def init_db():
        """Initialize Excel file"""
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Employees"
            ws.append(["Timestamp", "Action", "Name", "Wallet", "Telegram", "Language"])
            wb.save(EXCEL_FILE)
            logger.info(f"Created {EXCEL_FILE}")
    
    def save_employee(action: str, name: str, wallet: str, telegram: str, language: str):
        """Save employee data to Excel"""
        init_db()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            action, name, wallet, telegram, language
        ])
        wb.save(EXCEL_FILE)
        logger.info(f"Saved to Excel: {name}")
        return True
    
    def get_all_employees():
        """Get all employees from Excel"""
        if not os.path.exists(EXCEL_FILE):
            return []
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({
                "timestamp": row[0],
                "action": row[1],
                "name": row[2],
                "wallet": row[3],
                "telegram": row[4],
                "language": row[5]
            })
        return rows
